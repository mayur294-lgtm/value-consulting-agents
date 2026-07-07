"""roi-excel-generator component evaluator (objective, code-only — no LLM).

Verifies that `tools/roi_excel_generator.py` actually RENDERS the provenance
contract emitted by the roi-financial-modeler agent (tickets #83/#84), not
just that the agent's JSON contains the right keys. Each check GENERATES a
real .xlsx from a golden config, reopens it with openpyxl, and inspects the
actual cells/sheets:

  - sources_sheet_present / sources_sheet_absent_when_unset
        the "Sources" sheet renders IFF config['sources'] is set (additive,
        backward-compatible — evals/goldens/roi_config_no_provenance.json is
        the witness that old configs without provenance still generate clean).
  - no_confidence_row_leak
        the `_source` / `_confidence` companion keys must never leak into
        "Model Inputs" as their OWN input row (they're metadata on a field,
        not fields themselves).
  - explicit_confidence_wins
        a field's own `<field>_confidence` (e.g. HIGH from a client-confirmed
        source) must win over the TRANSCRIPT/BACKBASE keyword heuristic.
  - operating_costs_is_formula
        when the annual_revenue / cost_to_income_ratio / operating_costs trio
        is present, Operating Costs must render as a LIVE formula referencing
        the other two cells (not a hardcoded number).
  - derived_input_formula_renders
        a lever-driver input carrying a `formula` key must render as a LIVE
        Excel formula (with `{token}` placeholders substituted for real cell
        refs), not literal text.

target: path to a roi_config.json golden fixture (self-contained; no
engagement-output path resolution at runtime).

NOTE ON THE PAIRED FIXTURE: registry.yaml's `components:` altitude only wires
ONE `input:` target per component (no `goldens:`/`negatives:` list the way
deliverables get). Five of the six checks here are about the PROVENANCE
fixture (evals/goldens/roi_config_provenance.json, the wired `input:`). The
sixth — sources_sheet_absent_when_unset, the backward-compat witness — is
inherently about the OTHER fixture (evals/goldens/roi_config_no_provenance.json,
a config with no `sources` key at all). Rather than restructure run_experiment.py
to support multi-golden components, that one check loads its own sibling
fixture directly (same pattern as a judge check loading its own snapshot file
independently of `target`) so both fixtures are actually exercised by a single
`--component roi-excel-generator` run.
"""
from __future__ import annotations

import json
import sys
import tempfile
from pathlib import Path

from rubrics.base import CheckResult, repo_root


def _bool_check(name: str, ok: bool, *, hard_fail: bool = False, soft_floor: float = 0.0,
               detail: str = "") -> CheckResult:
    return CheckResult(name, 1.0 if ok else soft_floor, ok, hard_fail=hard_fail, detail=detail)


def _generate(config: dict) -> tuple[object | None, str | None]:
    """Run the real generator on `config`. Returns (workbook, error) — never raises.
    A generator crash is a CHECK FAILURE (not an eval-harness crash), so every
    check that depends on generation guards on this and fails clean if it errored."""
    root = repo_root()
    if str(root) not in sys.path:
        sys.path.insert(0, str(root))
    try:
        from tools.roi_excel_generator import ROIModelGenerator
        from openpyxl import load_workbook
    except ImportError as e:
        return None, f"import failed: {e}"
    try:
        with tempfile.TemporaryDirectory() as td:
            out = str(Path(td) / "eval_output.xlsx")
            ROIModelGenerator(config).generate(out)
            wb = load_workbook(out)
            return wb, None
    except Exception as e:  # noqa: BLE001 - a generator crash is a clean check failure
        return None, f"generation raised: {e}"


def _load_config(target: str) -> tuple[dict | None, str | None]:
    p = Path(target)
    if not p.exists():
        return None, f"fixture not found: {target}"
    try:
        return json.loads(p.read_text()), None
    except json.JSONDecodeError as e:
        return None, f"invalid JSON: {e}"


def _model_inputs_labels(wb) -> list[str]:
    ws = wb["Model Inputs"]
    labels = []
    for row in ws.iter_rows(min_col=2, max_col=2, values_only=True):
        v = row[0]
        if isinstance(v, str) and v.strip():
            labels.append(v.strip())
    return labels


def _basic_info_rows(wb) -> list[tuple[str, object, object]]:
    """Return (label, value_cell_value, confidence_cell_value) for each row under
    BASIC INFORMATION until the next section header (FINANCIAL PARAMETERS)."""
    ws = wb["Model Inputs"]
    rows = []
    in_section = False
    for row in ws.iter_rows(min_col=2, max_col=4, values_only=True):
        label, value, conf = row
        if label == "BASIC INFORMATION":
            in_section = True
            continue
        if label == "FINANCIAL PARAMETERS":
            break
        if in_section and isinstance(label, str) and label.strip():
            rows.append((label.strip(), value, conf))
    return rows


def _lever_driver_formula_cells(wb) -> list[str]:
    """Value cells under LEVER GROUP INPUTS that are LIVE formulas — i.e. a driver
    input whose `formula` key was rendered with its `{token}`s substituted for real
    cell refs (excludes the unrelated INDEX()/IF() scaffolding formulas elsewhere
    on the sheet, and excludes the Basic-Information Operating Costs formula)."""
    ws = wb["Model Inputs"]
    in_section = False
    out: list[str] = []
    for row in ws.iter_rows(min_col=2, max_col=3, values_only=True):
        label, value = row
        if label == "LEVER GROUP INPUTS":
            in_section = True
            continue
        if label == "ACTIVE BACKBASE IMPACTS (auto-selected from scenario)":
            break
        if in_section and isinstance(value, str) and value.startswith("=") \
                and not value.startswith("=INDEX(") and not value.startswith("=IF("):
            out.append(value)
    return out


def _check_sources_sheet_present(config: dict) -> CheckResult:
    wb, err = _generate(config)
    if err:
        return _bool_check("sources_sheet_present", False, detail=err)
    present = "Sources" in wb.sheetnames
    rows = 0
    if present:
        ws = wb["Sources"]
        # data rows start at row 6 (header row 5); count non-empty ref cells (col B).
        rows = sum(1 for r in ws.iter_rows(min_row=6, min_col=2, max_col=2, values_only=True) if r[0])
    ok = present and rows >= 1
    return _bool_check("sources_sheet_present", ok,
                       detail=f"Sources sheet present={present}, data rows={rows}")


_NO_PROVENANCE_GOLDEN = "evals/goldens/roi_config_no_provenance.json"


def _check_sources_sheet_absent_when_unset(_config: dict) -> CheckResult:
    """Backward-compat witness — always checked against the NO-PROVENANCE golden
    (not `target`/`_config`), since the registry only wires one `input:` slot and
    this behavior is specifically about the config WITHOUT a `sources` key."""
    root = repo_root()
    no_prov_config, load_err = _load_config(str(root / _NO_PROVENANCE_GOLDEN))
    if load_err:
        return _bool_check("sources_sheet_absent_when_unset", False, detail=load_err)
    wb, err = _generate(no_prov_config)
    if err:
        return _bool_check("sources_sheet_absent_when_unset", False, detail=err)
    absent = "Sources" not in wb.sheetnames
    return _bool_check("sources_sheet_absent_when_unset", absent,
                       detail="no_provenance golden: no Sources sheet (backward-compat, no exception)" if absent
                       else "no_provenance golden: unexpected Sources sheet with no 'sources' in config")


def _check_no_confidence_row_leak(config: dict) -> CheckResult:
    wb, err = _generate(config)
    if err:
        return _bool_check("no_confidence_row_leak", False, detail=err)
    labels = _model_inputs_labels(wb)
    leaked = [l for l in labels if l.endswith("Confidence")]
    ok = len(leaked) == 0
    return _bool_check("no_confidence_row_leak", ok,
                       detail="no '*_confidence' companion rendered as its own row" if ok
                       else f"leaked row label(s): {leaked}")


def _check_explicit_confidence_wins(config: dict) -> CheckResult:
    wb, err = _generate(config)
    if err:
        return _bool_check("explicit_confidence_wins", False, detail=err)
    for label, _value, conf in _basic_info_rows(wb):
        if label.lower() == "total fte":
            ok = conf == "HIGH"
            return _bool_check("explicit_confidence_wins", ok,
                               detail=f"Total Fte confidence cell = {conf!r} (expected HIGH)")
    return _bool_check("explicit_confidence_wins", False, detail="Total Fte row not found in Model Inputs")


def _check_operating_costs_is_formula(config: dict) -> CheckResult:
    wb, err = _generate(config)
    if err:
        return _bool_check("operating_costs_is_formula", False, detail=err)
    for label, value, _conf in _basic_info_rows(wb):
        if label.lower() == "operating costs":
            is_formula = isinstance(value, str) and value.startswith("=")
            return _bool_check("operating_costs_is_formula", is_formula,
                               detail=f"Operating Costs cell = {value!r}")
    return _bool_check("operating_costs_is_formula", False, detail="Operating Costs row not found in Model Inputs")


def _check_derived_input_formula_renders(config: dict) -> CheckResult:
    wb, err = _generate(config)
    if err:
        return _bool_check("derived_input_formula_renders", False, detail=err)
    formula_cells = _lever_driver_formula_cells(wb)
    ok = len(formula_cells) >= 1
    return _bool_check("derived_input_formula_renders", ok,
                       detail=f"{len(formula_cells)} derived-driver formula cell(s) found"
                       + (f" e.g. {formula_cells[0]!r}" if formula_cells else ""))


CHECKS = {
    "sources_sheet_present": _check_sources_sheet_present,
    "sources_sheet_absent_when_unset": _check_sources_sheet_absent_when_unset,
    "no_confidence_row_leak": _check_no_confidence_row_leak,
    "explicit_confidence_wins": _check_explicit_confidence_wins,
    "operating_costs_is_formula": _check_operating_costs_is_formula,
    "derived_input_formula_renders": _check_derived_input_formula_renders,
}


def evaluate(target: str) -> list[CheckResult]:
    """target: path to a roi_config.json golden fixture. Runs ALL registered
    checks against it (registry.yaml's `code:` list picks the relevant subset
    per fixture by calling run_experiment.py once per golden — here we just
    run everything; irrelevant checks for a given fixture still report a
    clean, informative pass/fail rather than being silently skipped)."""
    config, err = _load_config(target)
    if err:
        return [CheckResult(name, 0.0, False, hard_fail=True, detail=err) for name in CHECKS]
    return [fn(config) for fn in CHECKS.values()]
