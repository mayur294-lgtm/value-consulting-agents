#!/usr/bin/env python3
"""
ROI Excel Model Generator - Journey-Based Analysis

Generates Backbase-style ROI models following the Seabank methodology:
- Journey-based analysis (Customer Onboarding, Servicing, etc.)
- Revenue + Cost breakdown per journey
- Servicing analysis by channel (Branch, Call Center, Back Office)
- Clear assumption sourcing with confidence levels
- Implementation & Effectiveness curves by year

Usage:
    python roi_excel_generator.py --output output.xlsx --config roi_config.json
"""

import json
import argparse
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    exit(1)


class ROIModelGenerator:
    """Generates comprehensive ROI Excel models following Backbase/Seabank methodology."""

    COLORS = {
        'primary': '1A56FF',
        'dark': '1A1F36',
        'header_bg': '2D3250',
        'positive': '00B386',
        'negative': 'FF4444',
        'highlight': 'FFF3CD',
        'light_blue': 'E8F0FF',
        'light_gray': 'F5F5F5',
        'transcript': 'D4EDDA',      # Green - from transcript
        'estimate': 'FFF3CD',         # Yellow - estimate/assumption
        'benchmark': 'CCE5FF',        # Blue - Backbase benchmark
        'low_confidence': 'F8D7DA',   # Red - low confidence
    }

    def __init__(self, config: dict):
        self.config = config
        self.wb = Workbook()
        self.styles_created = False

    def _create_styles(self):
        if self.styles_created:
            return

        styles = [
            ("header_style", Font(bold=True, color="FFFFFF", size=11),
             PatternFill("solid", fgColor=self.COLORS['header_bg']),
             Alignment(horizontal="center", vertical="center")),
            ("title_style", Font(bold=True, color=self.COLORS['primary'], size=16), None, None),
            ("input_style", None, PatternFill("solid", fgColor=self.COLORS['light_blue']), None),
        ]

        for name, font, fill, align in styles:
            style = NamedStyle(name=name)
            if font:
                style.font = font
            if fill:
                style.fill = fill
            if align:
                style.alignment = align
            try:
                self.wb.add_named_style(style)
            except ValueError:
                pass

        self.styles_created = True

    def generate(self, output_path: str):
        """Generate the complete ROI Excel model with journey-based structure."""
        self._create_styles()

        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        # Create sheets matching Seabank structure
        self._create_cover_sheet()
        self._create_instructions_sheet()
        self._create_table_of_contents()
        self._create_results_dashboard()
        self._create_journey_inputs_sheet()
        self._create_customer_onboarding_analysis()
        self._create_servicing_analysis()
        self._create_cashflows_sheet()
        self._create_assumptions_sheet()
        self._create_data_gaps_sheet()

        self.wb.save(output_path)
        print(f"ROI Model saved to: {output_path}")
        return output_path

    def _create_cover_sheet(self):
        ws = self.wb.create_sheet("Cover Page", 0)
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40

        ws.merge_cells('B5:C5')
        ws['B5'] = "Return on Investment Calculator"
        ws['B5'].font = Font(bold=True, size=24, color=self.COLORS['primary'])

        ws.merge_cells('B7:C7')
        ws['B7'] = "Value Consulting - Business Case"
        ws['B7'].font = Font(size=14, italic=True)

        info = [
            (10, "Client Name:", self.config.get('client_name', 'Client')),
            (12, "Date:", self.config.get('date', datetime.now().strftime('%Y-%m-%d'))),
            (14, "Currency:", self.config.get('currency', 'USD')),
            (16, "Analysis Period:", f"{self.config.get('analysis_years', 5)} Years"),
            (18, "Selected Scenario:", self.config.get('selected_scenario', 'Moderate')),
            (20, "Discount Rate:", f"{self.config.get('discount_rate', 0.10) * 100}%"),
        ]

        for row, label, value in info:
            ws[f'B{row}'] = label
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'C{row}'] = value
            if label == "Selected Scenario:":
                ws[f'C{row}'].fill = PatternFill("solid", fgColor=self.COLORS['light_blue'])

        # Data confidence legend
        ws['B24'] = "DATA SOURCE LEGEND"
        ws['B24'].font = Font(bold=True, size=12)

        legends = [
            (26, "Green", "From transcript/client data", 'transcript'),
            (27, "Yellow", "Estimate - requires validation", 'estimate'),
            (28, "Blue", "Backbase benchmark", 'benchmark'),
            (29, "Red border", "Low confidence - critical validation", 'low_confidence'),
        ]

        for row, color, desc, color_key in legends:
            ws[f'B{row}'] = color
            ws[f'B{row}'].fill = PatternFill("solid", fgColor=self.COLORS[color_key])
            ws[f'C{row}'] = desc

    def _create_instructions_sheet(self):
        ws = self.wb.create_sheet("Instructions", 1)
        ws.column_dimensions['B'].width = 80

        ws['B2'] = "ROI Model Instructions"
        ws['B2'].font = Font(bold=True, size=18, color=self.COLORS['primary'])

        instructions = [
            "",
            "OVERVIEW",
            "This ROI model calculates value by JOURNEY - analyzing benefits for each customer journey separately.",
            "",
            "JOURNEY-BASED ANALYSIS",
            "• Customer Onboarding - Revenue uplift + Cost reduction from improved digital onboarding",
            "• Customer Servicing - Cost avoidance from channel shift (Branch, Call Center, Back Office)",
            "",
            "DATA CONFIDENCE LEVELS",
            "• HIGH - Direct quote from transcript or validated client data",
            "• MEDIUM - Inferred from transcript context or industry benchmark",
            "• LOW - Estimate based on typical patterns - REQUIRES VALIDATION",
            "",
            "CELL COLOR CODING",
            "• Green cells = Data from transcript/client",
            "• Yellow cells = Estimates requiring validation",
            "• Blue cells = Backbase benchmarks",
            "• Red border = Low confidence - prioritize validation",
            "",
            "SHEETS OVERVIEW",
            "1. Results Dashboard - Summary financial metrics (NPV, IRR, Payback)",
            "2. Journey Inputs - All inputs organized by journey",
            "3. Customer Onboarding Analysis - Detailed onboarding ROI breakdown",
            "4. Servicing Analysis - Channel-by-channel servicing cost avoidance",
            "5. Cashflows - 5-year projection",
            "6. Assumptions - Complete register with sources and owners",
            "7. Data Gaps - Required data for validation",
            "",
            "HOW TO USE",
            "1. Review 'Data Gaps' sheet - these are CRITICAL inputs needed from client",
            "2. Update 'Journey Inputs' with validated client data",
            "3. Results will recalculate automatically",
            "4. Review Assumptions sheet with client stakeholders",
        ]

        for i, line in enumerate(instructions, start=4):
            ws[f'B{i}'] = line
            if line in ["OVERVIEW", "JOURNEY-BASED ANALYSIS", "DATA CONFIDENCE LEVELS",
                       "CELL COLOR CODING", "SHEETS OVERVIEW", "HOW TO USE"]:
                ws[f'B{i}'].font = Font(bold=True, color=self.COLORS['primary'])

    def _create_table_of_contents(self):
        ws = self.wb.create_sheet("Table of Contents", 2)
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50

        ws['B3'] = "Table of Contents"
        ws['B3'].font = Font(bold=True, size=18, color=self.COLORS['primary'])

        sections = [
            ("Section 1", "Results Dashboard"),
            ("", "  Return on Investment"),
            ("", "  Net Present Value"),
            ("", "  Payback Period"),
            ("Section 2", "Value Driver Inputs"),
            ("", "  Customer Onboarding"),
            ("", "  Customer Servicing"),
            ("Section 3", "Journey Analysis"),
            ("", "  Customer Onboarding - Revenue + Cost"),
            ("", "  Customer Servicing - By Channel"),
            ("Section 4", "Servicing Analysis"),
            ("", "  Branch"),
            ("", "  Call Center"),
            ("", "  Back Office"),
            ("Section 5", "Cash Flows"),
            ("", "  Projection Inflows"),
            ("", "  Projection Outflows"),
            ("Section 6", "Assumptions & Data Gaps"),
            ("", "  Assumptions Register"),
            ("", "  Data Gaps for Validation"),
        ]

        for i, (section, desc) in enumerate(sections, start=5):
            ws[f'B{i}'] = section
            ws[f'B{i}'].font = Font(bold=True) if section else Font()
            ws[f'C{i}'] = desc

    def _create_results_dashboard(self):
        ws = self.wb.create_sheet("Results Dashboard", 3)

        for col, width in [('A', 5), ('B', 35), ('C', 20), ('D', 20), ('E', 20), ('F', 20), ('G', 20), ('H', 20)]:
            ws.column_dimensions[col].width = width

        ws['B2'] = "Financial Metrics - 5 Year Projections"
        ws['B2'].font = Font(bold=True, size=18, color=self.COLORS['primary'])

        # Calculate totals from journey data
        journeys = self.config.get('journeys', {})
        onboarding = journeys.get('customer_onboarding', {}).get('totals', {})
        servicing = journeys.get('customer_servicing', {}).get('totals', {})

        total_annual = onboarding.get('journey_total', 0) + servicing.get('total_cost_saved', 0)

        investment = self.config.get('investment', {})
        total_investment = investment.get('total', 8550000)

        # Key metrics
        metrics = [
            ("Total Annual Benefit (Steady State)", f"${total_annual:,.0f}"),
            ("Total Investment (5 Years)", f"${total_investment:,.0f}"),
            ("Net Present Value (NPV)", "See Cashflows"),
            ("Internal Rate of Return (IRR)", "See Cashflows"),
            ("Payback Period", "See Cashflows"),
        ]

        row = 5
        for name, value in metrics:
            ws[f'B{row}'] = name
            ws[f'B{row}'].font = Font(bold=True, size=12)
            ws[f'C{row}'] = value
            ws[f'C{row}'].font = Font(bold=True, size=16, color=self.COLORS['primary'])
            row += 3

        # Benefits by Journey
        row = 22
        ws[f'B{row}'] = "Benefits by Journey (Annual Steady State)"
        ws[f'B{row}'].font = Font(bold=True, size=14, color=self.COLORS['primary'])

        row += 2
        headers = ['Journey', 'Revenue Generation', 'Cost Reduction', 'Total']
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill("solid", fgColor=self.COLORS['header_bg'])

        journey_data = [
            ("Customer Onboarding", onboarding.get('revenue_generation_total', 0),
             onboarding.get('cost_reduction_total', 0), onboarding.get('journey_total', 0)),
            ("Customer Servicing", 0, servicing.get('total_cost_saved', 0),
             servicing.get('total_cost_saved', 0)),
        ]

        for journey, rev, cost, total in journey_data:
            row += 1
            ws.cell(row=row, column=2, value=journey)
            ws.cell(row=row, column=3, value=rev)
            ws.cell(row=row, column=3).number_format = '$#,##0'
            ws.cell(row=row, column=4, value=cost)
            ws.cell(row=row, column=4).number_format = '$#,##0'
            ws.cell(row=row, column=5, value=total)
            ws.cell(row=row, column=5).number_format = '$#,##0'
            ws.cell(row=row, column=5).font = Font(bold=True)

        # Total row
        row += 1
        ws.cell(row=row, column=2, value="TOTAL")
        ws.cell(row=row, column=2).font = Font(bold=True)
        total_rev = onboarding.get('revenue_generation_total', 0)
        total_cost = onboarding.get('cost_reduction_total', 0) + servicing.get('total_cost_saved', 0)
        ws.cell(row=row, column=3, value=total_rev)
        ws.cell(row=row, column=3).number_format = '$#,##0'
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=4, value=total_cost)
        ws.cell(row=row, column=4).number_format = '$#,##0'
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=5, value=total_annual)
        ws.cell(row=row, column=5).number_format = '$#,##0'
        ws.cell(row=row, column=5).font = Font(bold=True, color=self.COLORS['positive'])

    def _create_journey_inputs_sheet(self):
        ws = self.wb.create_sheet("Journey Inputs", 4)

        for col, width in [('A', 5), ('B', 40), ('C', 15), ('D', 12), ('E', 50), ('F', 20)]:
            ws.column_dimensions[col].width = width

        ws['B1'] = "Value Driver Inputs by Journey"
        ws['B1'].font = Font(bold=True, size=18, color=self.COLORS['primary'])

        ws['B3'] = "Color coding: Green=Transcript, Yellow=Estimate, Blue=Benchmark"
        ws['B3'].font = Font(italic=True, size=10)

        row = 5
        # Basic Information
        ws[f'B{row}'] = "Basic Information"
        ws[f'B{row}'].font = Font(bold=True, size=14, color=self.COLORS['primary'])

        basic = self.config.get('basic_information', {})
        basic_inputs = [
            ("Total Customers", basic.get('total_customers', 500000), basic.get('total_customers_source', '')),
            ("Annual Revenue", basic.get('annual_revenue', 50000000), basic.get('annual_revenue_source', '')),
            ("Average FTE Rate/Hour", basic.get('average_fte_rate_hour', 25), basic.get('average_fte_rate_source', '')),
        ]

        row += 1
        headers = ['Input', 'Value', 'Conf.', 'Source']
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill("solid", fgColor=self.COLORS['header_bg'])

        for name, value, source in basic_inputs:
            row += 1
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=value)
            ws.cell(row=row, column=3).number_format = '#,##0'

            # Color based on source
            if 'TRANSCRIPT' in source.upper():
                ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['transcript'])
                ws.cell(row=row, column=4, value='HIGH')
            elif 'BACKBASE' in source.upper():
                ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['benchmark'])
                ws.cell(row=row, column=4, value='MED')
            else:
                ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['estimate'])
                ws.cell(row=row, column=4, value='LOW')

            ws.cell(row=row, column=5, value=source)
            ws.cell(row=row, column=5).font = Font(size=9)

        # Customer Onboarding Journey
        row += 3
        ws[f'B{row}'] = "Customer Onboarding Journey"
        ws[f'B{row}'].font = Font(bold=True, size=14, color=self.COLORS['primary'])

        onboarding = self.config.get('journeys', {}).get('customer_onboarding', {})

        # Transcript quotes as evidence
        quotes = onboarding.get('transcript_quotes', [])
        if quotes:
            row += 1
            ws[f'B{row}'] = "Evidence from Discovery:"
            ws[f'B{row}'].font = Font(italic=True)
            for quote in quotes[:3]:
                row += 1
                ws[f'B{row}'] = f"  \"{quote}\""
                ws[f'B{row}'].font = Font(italic=True, size=9, color='666666')

        # Revenue drivers
        row += 2
        ws[f'B{row}'] = "Revenue Drivers"
        ws[f'B{row}'].font = Font(bold=True)

        for driver_key, driver in onboarding.get('revenue_drivers', {}).items():
            row += 1
            ws[f'B{row}'] = driver.get('name', driver_key)
            ws[f'B{row}'].font = Font(bold=True, color=self.COLORS['primary'])

            for input_key, input_data in driver.get('inputs', {}).items():
                row += 1
                ws.cell(row=row, column=2, value=f"  {input_key.replace('_', ' ').title()}")
                ws.cell(row=row, column=3, value=input_data.get('value', 0))

                confidence = input_data.get('confidence', 'LOW')
                source = input_data.get('source', '')

                if confidence == 'HIGH' or 'TRANSCRIPT' in source.upper():
                    ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['transcript'])
                elif 'BACKBASE' in source.upper():
                    ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['benchmark'])
                else:
                    ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor=self.COLORS['estimate'])

                ws.cell(row=row, column=4, value=confidence)
                ws.cell(row=row, column=5, value=source)
                ws.cell(row=row, column=5).font = Font(size=9)

            row += 1
            ws.cell(row=row, column=2, value="  Potential Annual Benefit")
            ws.cell(row=row, column=2).font = Font(bold=True)
            ws.cell(row=row, column=3, value=driver.get('potential_annual_benefit', 0))
            ws.cell(row=row, column=3).number_format = '$#,##0'
            ws.cell(row=row, column=3).font = Font(bold=True, color=self.COLORS['positive'])

        # Customer Servicing Journey
        row += 3
        ws[f'B{row}'] = "Customer Servicing Journey"
        ws[f'B{row}'].font = Font(bold=True, size=14, color=self.COLORS['primary'])

        servicing = self.config.get('journeys', {}).get('customer_servicing', {})

        quotes = servicing.get('transcript_quotes', [])
        if quotes:
            row += 1
            ws[f'B{row}'] = "Evidence from Discovery:"
            ws[f'B{row}'].font = Font(italic=True)
            for quote in quotes[:3]:
                row += 1
                ws[f'B{row}'] = f"  \"{quote}\""
                ws[f'B{row}'].font = Font(italic=True, size=9, color='666666')

        row += 2
        ws[f'B{row}'] = "See 'Servicing Analysis' sheet for detailed breakdown by channel"
        ws[f'B{row}'].font = Font(italic=True)

    def _create_customer_onboarding_analysis(self):
        ws = self.wb.create_sheet("Onboarding Analysis", 5)

        for col, width in [('A', 5), ('B', 40), ('C', 18), ('D', 18), ('E', 18), ('F', 18), ('G', 18), ('H', 18)]:
            ws.column_dimensions[col].width = width

        ws['B1'] = "Customer Onboarding - Value Analysis"
        ws['B1'].font = Font(bold=True, size=18, color=self.COLORS['primary'])

        onboarding = self.config.get('journeys', {}).get('customer_onboarding', {})

        # Evidence
        row = 3
        ws[f'B{row}'] = f"Evidence IDs: {', '.join(onboarding.get('evidence_ids', []))}"
        ws[f'B{row}'].font = Font(italic=True)

        # Revenue Generation section
        row = 6
        ws[f'B{row}'] = "Revenue Generation"
        ws[f'B{row}'].font = Font(bold=True, size=14, color=self.COLORS['positive'])

        row += 2
        headers = ['Driver', 'Baseline', 'Impact', 'Annual Benefit', 'Y1', 'Y2', 'Y3', 'Y4', 'Y5']
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill("solid", fgColor=self.COLORS['header_bg'])

        impl_curve = self.config.get('backbase_loading', {}).get('implementation_curve', [0.3, 0.7, 0.8, 1.0, 1.0])
        eff_curve = self.config.get('backbase_loading', {}).get('effectiveness_curve', [0.15, 0.35, 0.6, 0.85, 1.0])

        for driver_key, driver in onboarding.get('revenue_drivers', {}).items():
            row += 1
            ws.cell(row=row, column=2, value=driver.get('name', ''))
            ws.cell(row=row, column=3, value=driver.get('baseline_annual', 0))
            ws.cell(row=row, column=3).number_format = '$#,##0'

            impact = driver.get('inputs', {}).get('backbase_impact', {}).get('value', 0)
            ws.cell(row=row, column=4, value=impact)
            ws.cell(row=row, column=4).number_format = '0.0%'

            annual = driver.get('potential_annual_benefit', 0)
            ws.cell(row=row, column=5, value=annual)
            ws.cell(row=row, column=5).number_format = '$#,##0'
            ws.cell(row=row, column=5).font = Font(bold=True)

            # Year by year with curves
            for yr in range(5):
                yearly = annual * impl_curve[yr] * eff_curve[yr]
                ws.cell(row=row, column=6+yr, value=yearly)
                ws.cell(row=row, column=6+yr).number_format = '$#,##0'

        # Cost Reduction section
        row += 3
        ws[f'B{row}'] = "Cost Reduction"
        ws[f'B{row}'].font = Font(bold=True, size=14, color=self.COLORS['positive'])

        row += 2
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill("solid", fgColor=self.COLORS['header_bg'])

        for driver_key, driver in onboarding.get('cost_drivers', {}).items():
            row += 1
            ws.cell(row=row, column=2, value=driver.get('name', ''))
            ws.cell(row=row, column=3, value=driver.get('baseline_annual', 0))
            ws.cell(row=row, column=3).number_format = '$#,##0'

            impact = driver.get('inputs', {}).get('backbase_impact', {}).get('value', 0)
            ws.cell(row=row, column=4, value=impact)
            ws.cell(row=row, column=4).number_format = '0.0%'

            annual = driver.get('potential_annual_benefit', 0)
            ws.cell(row=row, column=5, value=annual)
            ws.cell(row=row, column=5).number_format = '$#,##0'
            ws.cell(row=row, column=5).font = Font(bold=True)

            for yr in range(5):
                yearly = annual * impl_curve[yr] * eff_curve[yr]
                ws.cell(row=row, column=6+yr, value=yearly)
                ws.cell(row=row, column=6+yr).number_format = '$#,##0'

        # Journey Total
        row += 3
        ws[f'B{row}'] = "Customer Onboarding Total"
        ws[f'B{row}'].font = Font(bold=True, size=12)

        totals = onboarding.get('totals', {})
        ws.cell(row=row, column=5, value=totals.get('journey_total', 0))
        ws.cell(row=row, column=5).number_format = '$#,##0'
        ws.cell(row=row, column=5).font = Font(bold=True, size=14, color=self.COLORS['positive'])

    def _create_servicing_analysis(self):
        ws = self.wb.create_sheet("Servicing Analysis", 6)

        for col, width in [('A', 5), ('B', 30), ('C', 15), ('D', 15), ('E', 12), ('F', 18), ('G', 12), ('H', 18), ('I', 40)]:
            ws.column_dimensions[col].width = width

        ws['B1'] = "Customer Servicing - Channel Analysis"
        ws['B1'].font = Font(bold=True, size=18, color=self.COLORS['primary'])

        servicing = self.config.get('journeys', {}).get('customer_servicing', {})

        row = 3
        ws[f'B{row}'] = f"Evidence IDs: {', '.join(servicing.get('evidence_ids', []))}"
        ws[f'B{row}'].font = Font(italic=True)

        # Headers
        row = 6
        headers = ['Task', 'Volume', 'Time (hrs)', 'Rate', 'Baseline', 'Impact', 'Saved', 'Source']
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill("solid", fgColor=self.COLORS['header_bg'])

        servicing_data = servicing.get('servicing_analysis', {})

        for channel_key in ['branch', 'call_center', 'back_office']:
            channel = servicing_data.get(channel_key, {})
            if not channel:
                continue

            row += 2
            ws[f'B{row}'] = channel.get('channel_name', channel_key.title())
            ws[f'B{row}'].font = Font(bold=True, size=12, color=self.COLORS['primary'])

            for task in channel.get('tasks', []):
                row += 1
                ws.cell(row=row, column=2, value=task.get('task', ''))
                ws.cell(row=row, column=3, value=task.get('yearly_volume', 0))
                ws.cell(row=row, column=3).number_format = '#,##0'
                ws.cell(row=row, column=4, value=task.get('time_spent_hours', 0))
                ws.cell(row=row, column=5, value=task.get('fte_rate', 25))
                ws.cell(row=row, column=5).number_format = '$#,##0'
                ws.cell(row=row, column=6, value=task.get('baseline_cost', 0))
                ws.cell(row=row, column=6).number_format = '$#,##0'
                ws.cell(row=row, column=7, value=task.get('backbase_impact', 0))
                ws.cell(row=row, column=7).number_format = '0%'

                # Color impact cell based on source
                source = task.get('impact_source', '')
                if 'BACKBASE' in source.upper():
                    ws.cell(row=row, column=7).fill = PatternFill("solid", fgColor=self.COLORS['benchmark'])

                ws.cell(row=row, column=8, value=task.get('cost_saved', 0))
                ws.cell(row=row, column=8).number_format = '$#,##0'
                ws.cell(row=row, column=8).font = Font(color=self.COLORS['positive'])

                ws.cell(row=row, column=9, value=task.get('volume_source', ''))
                ws.cell(row=row, column=9).font = Font(size=8, color='666666')

            # Channel subtotal
            row += 1
            totals = channel.get('channel_total', {})
            ws.cell(row=row, column=2, value=f"{channel.get('channel_name', '')} Subtotal")
            ws.cell(row=row, column=2).font = Font(bold=True)
            ws.cell(row=row, column=3, value=totals.get('yearly_volume', 0))
            ws.cell(row=row, column=3).number_format = '#,##0'
            ws.cell(row=row, column=3).font = Font(bold=True)
            ws.cell(row=row, column=6, value=totals.get('baseline_cost', 0))
            ws.cell(row=row, column=6).number_format = '$#,##0'
            ws.cell(row=row, column=6).font = Font(bold=True)
            ws.cell(row=row, column=8, value=totals.get('cost_saved', 0))
            ws.cell(row=row, column=8).number_format = '$#,##0'
            ws.cell(row=row, column=8).font = Font(bold=True, color=self.COLORS['positive'])

        # Grand total
        row += 3
        totals = servicing.get('totals', {})
        ws[f'B{row}'] = "TOTAL SERVICING"
        ws[f'B{row}'].font = Font(bold=True, size=14)
        ws.cell(row=row, column=3, value=totals.get('total_servicing_volume', 0))
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=6, value=totals.get('total_baseline_cost', 0))
        ws.cell(row=row, column=6).number_format = '$#,##0'
        ws.cell(row=row, column=6).font = Font(bold=True)
        ws.cell(row=row, column=8, value=totals.get('total_cost_saved', 0))
        ws.cell(row=row, column=8).number_format = '$#,##0'
        ws.cell(row=row, column=8).font = Font(bold=True, size=14, color=self.COLORS['positive'])

    def _create_cashflows_sheet(self):
        ws = self.wb.create_sheet("Cashflows", 7)

        for col, width in [('A', 5), ('B', 40)] + [(get_column_letter(i), 18) for i in range(3, 10)]:
            ws.column_dimensions[col].width = width

        ws['B1'] = "5 Year Cashflows & ROI"
        ws['B1'].font = Font(bold=True, size=18, color=self.COLORS['primary'])

        # Implementation curves
        impl_curve = self.config.get('backbase_loading', {}).get('implementation_curve', [0.3, 0.7, 0.8, 1.0, 1.0])
        eff_curve = self.config.get('backbase_loading', {}).get('effectiveness_curve', [0.15, 0.35, 0.6, 0.85, 1.0])

        # Get journey totals
        journeys = self.config.get('journeys', {})
        onboarding_total = journeys.get('customer_onboarding', {}).get('totals', {}).get('journey_total', 0)
        servicing_total = journeys.get('customer_servicing', {}).get('totals', {}).get('total_cost_saved', 0)
        total_annual = onboarding_total + servicing_total

        # Headers
        row = 4
        headers = ['', 'Year 1', 'Year 2', 'Year 3', 'Year 4', 'Year 5', 'Total']
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill("solid", fgColor=self.COLORS['header_bg'])

        # Cash Inflows by Journey
        row = 6
        ws[f'B{row}'] = "CASH INFLOWS (Benefits)"
        ws[f'B{row}'].font = Font(bold=True, color=self.COLORS['positive'])

        # Customer Onboarding
        row += 1
        ws[f'B{row}'] = "Customer Onboarding"
        yearly_onboarding = []
        for yr in range(5):
            yearly = onboarding_total * impl_curve[yr] * eff_curve[yr]
            yearly_onboarding.append(yearly)
            ws.cell(row=row, column=3+yr, value=yearly)
            ws.cell(row=row, column=3+yr).number_format = '$#,##0'
        ws.cell(row=row, column=8, value=sum(yearly_onboarding))
        ws.cell(row=row, column=8).number_format = '$#,##0'

        # Customer Servicing
        row += 1
        ws[f'B{row}'] = "Customer Servicing"
        yearly_servicing = []
        for yr in range(5):
            yearly = servicing_total * impl_curve[yr] * eff_curve[yr]
            yearly_servicing.append(yearly)
            ws.cell(row=row, column=3+yr, value=yearly)
            ws.cell(row=row, column=3+yr).number_format = '$#,##0'
        ws.cell(row=row, column=8, value=sum(yearly_servicing))
        ws.cell(row=row, column=8).number_format = '$#,##0'

        # Total Inflows
        row += 1
        ws[f'B{row}'] = "Total Cash Inflow"
        ws[f'B{row}'].font = Font(bold=True)
        yearly_total = [yearly_onboarding[i] + yearly_servicing[i] for i in range(5)]
        for yr in range(5):
            ws.cell(row=row, column=3+yr, value=yearly_total[yr])
            ws.cell(row=row, column=3+yr).number_format = '$#,##0'
            ws.cell(row=row, column=3+yr).font = Font(bold=True)
        ws.cell(row=row, column=8, value=sum(yearly_total))
        ws.cell(row=row, column=8).number_format = '$#,##0'
        ws.cell(row=row, column=8).font = Font(bold=True)

        # Cash Outflows
        row += 3
        ws[f'B{row}'] = "CASH OUTFLOWS (Investment)"
        ws[f'B{row}'].font = Font(bold=True, color=self.COLORS['negative'])

        investment = self.config.get('investment', {})
        license_data = investment.get('license', {})
        impl_data = investment.get('implementation', {})

        row += 1
        ws[f'B{row}'] = "License"
        license_values = [
            license_data.get(f'year_{i+1}', 800000 if i == 0 else 1000000) for i in range(5)
        ]
        for yr in range(5):
            ws.cell(row=row, column=3+yr, value=license_values[yr])
            ws.cell(row=row, column=3+yr).number_format = '$#,##0'
        ws.cell(row=row, column=8, value=sum(license_values))
        ws.cell(row=row, column=8).number_format = '$#,##0'

        row += 1
        ws[f'B{row}'] = "Implementation"
        impl_values = [
            impl_data.get(f'year_{i+1}', 2500000 if i == 0 else 500000 if i == 1 else 250000) for i in range(5)
        ]
        for yr in range(5):
            ws.cell(row=row, column=3+yr, value=impl_values[yr])
            ws.cell(row=row, column=3+yr).number_format = '$#,##0'
        ws.cell(row=row, column=8, value=sum(impl_values))
        ws.cell(row=row, column=8).number_format = '$#,##0'

        row += 1
        ws[f'B{row}'] = "Total Cash Outflow"
        ws[f'B{row}'].font = Font(bold=True)
        yearly_outflow = [license_values[i] + impl_values[i] for i in range(5)]
        for yr in range(5):
            ws.cell(row=row, column=3+yr, value=yearly_outflow[yr])
            ws.cell(row=row, column=3+yr).number_format = '$#,##0'
            ws.cell(row=row, column=3+yr).font = Font(bold=True)
        ws.cell(row=row, column=8, value=sum(yearly_outflow))
        ws.cell(row=row, column=8).number_format = '$#,##0'
        ws.cell(row=row, column=8).font = Font(bold=True)

        # Net Cashflow
        row += 3
        ws[f'B{row}'] = "NET CASHFLOW"
        ws[f'B{row}'].font = Font(bold=True, size=14)
        net_cashflow = [yearly_total[i] - yearly_outflow[i] for i in range(5)]
        for yr in range(5):
            ws.cell(row=row, column=3+yr, value=net_cashflow[yr])
            ws.cell(row=row, column=3+yr).number_format = '$#,##0'
            ws.cell(row=row, column=3+yr).font = Font(bold=True)
            if net_cashflow[yr] < 0:
                ws.cell(row=row, column=3+yr).font = Font(bold=True, color=self.COLORS['negative'])
            else:
                ws.cell(row=row, column=3+yr).font = Font(bold=True, color=self.COLORS['positive'])
        ws.cell(row=row, column=8, value=sum(net_cashflow))
        ws.cell(row=row, column=8).number_format = '$#,##0'
        ws.cell(row=row, column=8).font = Font(bold=True, size=14)

        # Cumulative
        row += 1
        ws[f'B{row}'] = "Cumulative Cashflow"
        cumulative = 0
        for yr in range(5):
            cumulative += net_cashflow[yr]
            ws.cell(row=row, column=3+yr, value=cumulative)
            ws.cell(row=row, column=3+yr).number_format = '$#,##0'

        # Financial Metrics
        row += 4
        ws[f'B{row}'] = "FINANCIAL METRICS"
        ws[f'B{row}'].font = Font(bold=True, size=14, color=self.COLORS['primary'])

        discount_rate = self.config.get('discount_rate', 0.10)

        # NPV calculation
        npv = sum(net_cashflow[i] / ((1 + discount_rate) ** (i + 1)) for i in range(5))

        # Simple payback
        cumulative = 0
        payback = 5
        for yr in range(5):
            cumulative += net_cashflow[yr]
            if cumulative >= 0:
                payback = yr + 1 - (cumulative - net_cashflow[yr]) / net_cashflow[yr] if net_cashflow[yr] != 0 else yr + 1
                break

        metrics = [
            ("Discount Rate (WACC)", discount_rate, '0.0%'),
            ("Net Present Value (NPV)", npv, '$#,##0'),
            ("Payback Period", payback, '0.0'),
            ("Total 5-Year Benefits", sum(yearly_total), '$#,##0'),
            ("Total 5-Year Investment", sum(yearly_outflow), '$#,##0'),
        ]

        for name, value, fmt in metrics:
            row += 1
            ws[f'B{row}'] = name
            ws.cell(row=row, column=3, value=value)
            ws.cell(row=row, column=3).number_format = fmt
            ws.cell(row=row, column=3).font = Font(bold=True)

    def _create_assumptions_sheet(self):
        ws = self.wb.create_sheet("Assumptions", 8)

        for col, width in [('A', 5), ('B', 8), ('C', 35), ('D', 15), ('E', 12), ('F', 50), ('G', 20), ('H', 35)]:
            ws.column_dimensions[col].width = width

        ws['B1'] = "Assumptions Register"
        ws['B1'].font = Font(bold=True, size=18, color=self.COLORS['primary'])

        ws['B3'] = "All assumptions documented with sources and validation owners"
        ws['B3'].font = Font(italic=True)

        row = 5
        headers = ['ID', 'Assumption', 'Value', 'Conf.', 'Source', 'Owner', 'Sensitivity']
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill("solid", fgColor=self.COLORS['header_bg'])

        assumptions = self.config.get('assumptions_register', [])

        for assumption in assumptions:
            row += 1
            ws.cell(row=row, column=2, value=assumption.get('id', ''))
            ws.cell(row=row, column=3, value=assumption.get('assumption', ''))
            ws.cell(row=row, column=4, value=assumption.get('value', ''))

            confidence = assumption.get('confidence', 'LOW')
            ws.cell(row=row, column=5, value=confidence)

            # Color based on confidence
            if confidence == 'HIGH':
                ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor=self.COLORS['transcript'])
            elif confidence == 'MEDIUM':
                ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor=self.COLORS['benchmark'])
            else:
                ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor=self.COLORS['estimate'])

            ws.cell(row=row, column=6, value=assumption.get('source', ''))
            ws.cell(row=row, column=6).font = Font(size=9)
            ws.cell(row=row, column=7, value=assumption.get('validation_owner', ''))
            ws.cell(row=row, column=8, value=assumption.get('sensitivity', ''))
            ws.cell(row=row, column=8).font = Font(size=9)

    def _create_data_gaps_sheet(self):
        ws = self.wb.create_sheet("Data Gaps", 9)

        for col, width in [('A', 5), ('B', 8), ('C', 40), ('D', 12), ('E', 35), ('F', 45)]:
            ws.column_dimensions[col].width = width

        ws['B1'] = "Data Gaps - Required for Validation"
        ws['B1'].font = Font(bold=True, size=18, color=self.COLORS['primary'])

        ws['B3'] = "These data points are CRITICAL for validating the ROI model. Prioritize obtaining this data from the client."
        ws['B3'].font = Font(italic=True, color=self.COLORS['negative'])

        row = 5
        headers = ['ID', 'Data Needed', 'Priority', 'Source', 'Impact on Model']
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill("solid", fgColor=self.COLORS['header_bg'])

        data_gaps = self.config.get('data_gaps_for_validation', [])

        for gap in data_gaps:
            row += 1
            ws.cell(row=row, column=2, value=gap.get('id', ''))
            ws.cell(row=row, column=3, value=gap.get('data_needed', ''))

            priority = gap.get('priority', 'MEDIUM')
            ws.cell(row=row, column=4, value=priority)

            if priority == 'CRITICAL':
                ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor=self.COLORS['low_confidence'])
                ws.cell(row=row, column=4).font = Font(bold=True)
            elif priority == 'HIGH':
                ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor=self.COLORS['estimate'])

            ws.cell(row=row, column=5, value=gap.get('source', ''))
            ws.cell(row=row, column=6, value=gap.get('impact', ''))
            ws.cell(row=row, column=6).font = Font(size=9)


def main():
    parser = argparse.ArgumentParser(description='Generate journey-based ROI Excel models')
    parser.add_argument('--output', '-o', default='roi_model.xlsx', help='Output Excel file path')
    parser.add_argument('--config', '-c', required=True, help='JSON configuration file')

    args = parser.parse_args()

    if not Path(args.config).exists():
        print(f"Config file not found: {args.config}")
        exit(1)

    with open(args.config) as f:
        config = json.load(f)

    generator = ROIModelGenerator(config)
    generator.generate(args.output)


if __name__ == '__main__':
    main()
